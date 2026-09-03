"""The `crfs` worksheet reader.

This module had no tests at all. Every field is read by column *position*, and
nothing checked the headers -- so inserting a column shifted `displayname`
into `primarykey` and the idconfig JSON into `isbase`, producing a manifest
that was well-formed JSON and wrong for every form, which no later stage can
detect.
"""

import unittest

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from crf_reader import CRFS_COLUMN_NAMES, CrfReader


ENROLLEE_IDCONFIG = (
    '{"prefix":"","fields":[{"name":"deviceid","length":3},'
    '{"name":"mrc","length":3}],"incrementLength":4}'
)


def sheet(rows, headers=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "crfs"
    worksheet.append(list(CRFS_COLUMN_NAMES) if headers is None else headers)
    for r in rows:
        worksheet.append(r)
    return worksheet


def crf_row(
    display_order=10,
    tablename="enrollee",
    displayname="Enrollee",
    primarykey="subjid",
    idconfig="",
    isbase=1,
    **rest,
):
    """One crfs row in column order."""
    return [
        display_order, tablename, displayname, primarykey, idconfig, isbase,
        rest.get("linkingfield", ""), rest.get("parenttable", ""),
        rest.get("incrementfield", ""), rest.get("requireslink", ""),
        rest.get("repeat_count_field", ""), rest.get("auto_start_repeat", ""),
        rest.get("repeat_enforce_count", ""), rest.get("display_fields", ""),
        rest.get("entry_condition", ""),
    ]


class HeaderTests(unittest.TestCase):
    def test_the_documented_headers_are_accepted(self):
        crfs, errors = CrfReader.read_crfs_worksheet(sheet([crf_row()]))
        self.assertEqual([], errors)
        self.assertEqual(1, len(crfs))
        self.assertEqual("enrollee", crfs[0].tablename)

    def test_headers_are_matched_case_insensitively(self):
        headers = [h.upper() for h in CRFS_COLUMN_NAMES]
        crfs, errors = CrfReader.read_crfs_worksheet(sheet([crf_row()], headers))
        self.assertEqual([], errors)
        self.assertEqual(1, len(crfs))

    def test_an_inserted_column_is_an_error_not_a_silent_shift(self):
        """The bug this check exists for: a column added before `idconfig`."""
        headers = list(CRFS_COLUMN_NAMES)
        headers.insert(4, "notes")
        crfs, errors = CrfReader.read_crfs_worksheet(sheet([crf_row()], headers))
        self.assertTrue(errors)
        self.assertIn("header names", errors[0])
        # Nothing is returned: positions cannot be trusted, so reading rows
        # would invent data.
        self.assertEqual([], crfs)

    def test_reordered_columns_are_an_error(self):
        headers = list(CRFS_COLUMN_NAMES)
        headers[1], headers[2] = headers[2], headers[1]
        _, errors = CrfReader.read_crfs_worksheet(sheet([crf_row()], headers))
        self.assertTrue(errors)

    def test_the_error_names_what_was_expected_and_what_was_found(self):
        headers = list(CRFS_COLUMN_NAMES)
        headers[4] = "id_config"
        _, errors = CrfReader.read_crfs_worksheet(sheet([crf_row()], headers))
        self.assertIn("idconfig", errors[0])
        self.assertIn("id_config", errors[0])


class BlankRowTests(unittest.TestCase):
    def test_a_trailing_blank_row_is_not_a_form(self):
        """openpyxl counts a format-only row toward max_row.

        Such a row used to become an all-None Crf, which the manifest
        serialized as a bare `{}` entry in its crfs array.
        """
        worksheet = sheet([crf_row()])
        # A border and nothing else -- the real-world case. This is what
        # inflates max_row past the last row that holds any data.
        worksheet.cell(row=3, column=3).border = Border(bottom=Side(style="thin"))
        self.assertGreaterEqual(worksheet.max_row, 3)
        crfs, errors = CrfReader.read_crfs_worksheet(worksheet)
        self.assertEqual([], errors)
        self.assertEqual(1, len(crfs))

    def test_a_blank_row_between_forms_is_skipped(self):
        worksheet = sheet([crf_row(tablename="a"), [None] * 15, crf_row(tablename="b")])
        crfs, errors = CrfReader.read_crfs_worksheet(worksheet)
        self.assertEqual([], errors)
        self.assertEqual(["a", "b"], [c.tablename for c in crfs])


class IdConfigTests(unittest.TestCase):
    def test_a_valid_idconfig_is_parsed(self):
        crfs, errors = CrfReader.read_crfs_worksheet(
            sheet([crf_row(idconfig=ENROLLEE_IDCONFIG)])
        )
        self.assertEqual([], errors)
        idconfig = crfs[0].idconfig
        self.assertEqual(4, idconfig.incrementLength)
        self.assertEqual(["deviceid", "mrc"], [f.name for f in idconfig.fields])
        self.assertEqual([3, 3], [f.length for f in idconfig.fields])

    def test_a_blank_idconfig_is_simply_absent(self):
        crfs, errors = CrfReader.read_crfs_worksheet(sheet([crf_row(idconfig="")]))
        self.assertEqual([], errors)
        self.assertIsNone(crfs[0].idconfig)

    def test_malformed_idconfig_json_is_reported_not_swallowed(self):
        """It used to become `None` with no log line at all.

        idconfig builds the study's primary keys, so a silently dropped one
        means the app generates a different ID scheme than intended.
        """
        crfs, errors = CrfReader.read_crfs_worksheet(
            sheet([crf_row(idconfig='{"prefix":"", "fields":[}')])
        )
        self.assertTrue(errors)
        self.assertIn("idconfig", errors[0])
        self.assertIn("enrollee", errors[0])

    def test_a_smart_quote_in_idconfig_is_reported(self):
        crfs, errors = CrfReader.read_crfs_worksheet(
            sheet([crf_row(idconfig='{“prefix”:""}')])
        )
        self.assertTrue(errors)

    def test_idconfig_that_is_valid_json_but_not_an_object_is_reported(self):
        _, errors = CrfReader.read_crfs_worksheet(sheet([crf_row(idconfig="[1,2]")]))
        self.assertTrue(errors)

    def test_the_row_number_is_in_the_message(self):
        rows = [crf_row(tablename="a"), crf_row(tablename="b", idconfig="{bad")]
        _, errors = CrfReader.read_crfs_worksheet(sheet(rows))
        self.assertIn("Row 3", errors[0])


class ColumnMappingTests(unittest.TestCase):
    def test_every_column_lands_in_its_own_field(self):
        """Guards the positional mapping itself."""
        crfs, errors = CrfReader.read_crfs_worksheet(
            sheet([[
                20, "nets", "Nets", "netid", ENROLLEE_IDCONFIG, 0,
                "hhid", "hh_info", "netnum", 1,
                "numnets", 2, 3, "netid,netshape", "eligible=1",
            ]])
        )
        self.assertEqual([], errors)
        crf = crfs[0]
        self.assertEqual(20, crf.display_order)
        self.assertEqual("nets", crf.tablename)
        self.assertEqual("Nets", crf.displayname)
        self.assertEqual("netid", crf.primarykey)
        self.assertEqual(0, crf.isbase)
        self.assertEqual("hhid", crf.linkingfield)
        self.assertEqual("hh_info", crf.parenttable)
        self.assertEqual("netnum", crf.incrementfield)
        self.assertEqual(1, crf.requireslink)
        self.assertEqual("numnets", crf.repeat_count_field)
        self.assertEqual(2, crf.auto_start_repeat)
        self.assertEqual(3, crf.repeat_enforce_count)
        self.assertEqual("netid,netshape", crf.display_fields)
        self.assertEqual("eligible=1", crf.entry_condition)
        self.assertIsNotNone(crf.idconfig)


if __name__ == "__main__":
    unittest.main()
