"""What reaches disk when a build fails.

The gate used to be evaluated once, before the per-file XML validation inside
the loop could set `errorsEncountered` -- so a validation failure skipped only
the zip. Every .xml and survey_manifest.gistx stayed in outputPath while the
console printed "The XML files and manifest HAVE NOT been created", and
hand-zipping those files ships a package containing malformed XML.
"""

import io
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

from crf_reader import CRFS_COLUMN_NAMES
from models import AppConfig
from processor import SurveyGenProcessor

from tests.test_dd_validations import HEADERS, numeric_row, row


def workbook_file(directory, crfs_headers=None, crfs_rows=None):
    """A minimal but complete dictionary: one `_dd` sheet plus a `crfs` sheet."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "enrollee_dd"
    worksheet.append(HEADERS)
    # The crfs row below names `subjid` as the primary key, so the sheet has
    # to declare it -- both real dictionaries do. A primarykey naming no real
    # column means the app never creates it, so getAllPrimaryKeys fails and
    # the duplicate check silently does nothing; that is now a build error.
    worksheet.append(row("subjid", "automatic", "text", "Subject ID"))
    worksheet.append(numeric_row("age"))

    crfs = workbook.create_sheet("crfs")
    crfs.append(list(CRFS_COLUMN_NAMES) if crfs_headers is None else crfs_headers)
    for r in crfs_rows if crfs_rows is not None else [
        [10, "enrollee", "Enrollee", "subjid", "", 1, "", "", "", "", "", "", "", "", ""]
    ]:
        crfs.append(r)

    path = Path(directory) / "dictionary.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def run_quietly(processor):
    """`run` prints the Windows app's SUCCESS/ERRORS message box equivalent."""
    with redirect_stdout(io.StringIO()):
        return processor.run()


def config_for(excel_path, output_path):
    return AppConfig(
        excelFile=str(excel_path),
        csvFiles="",
        outputPath=str(output_path),
        surveyName="Test",
        surveyId="test_survey",
        databaseName="test.sqlite",
    )


class SuccessfulBuildTests(unittest.TestCase):
    def test_a_clean_dictionary_produces_only_a_zip_and_a_log(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            processor = SurveyGenProcessor(config_for(workbook_file(tmp), out))
            self.assertEqual(0, run_quietly(processor), "\n".join(processor.logstring))

            names = sorted(p.name for p in out.iterdir())
            self.assertEqual(["gistlogfile.txt", "test_survey.zip"], names)

    def test_the_zip_holds_the_xml_and_the_manifest(self):
        from zipfile import ZipFile

        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            run_quietly(SurveyGenProcessor(config_for(workbook_file(tmp), out)))
            with ZipFile(out / "test_survey.zip") as archive:
                self.assertEqual(
                    ["enrollee.xml", "survey_manifest.gistx"],
                    sorted(archive.namelist()),
                )

    def test_the_manifest_names_the_file_that_is_actually_in_the_zip(self):
        import json
        from zipfile import ZipFile

        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            run_quietly(SurveyGenProcessor(config_for(workbook_file(tmp), out)))
            with ZipFile(out / "test_survey.zip") as archive:
                manifest = json.loads(archive.read("survey_manifest.gistx"))
                members = set(archive.namelist())
            for xml_name in manifest["xmlFiles"]:
                self.assertIn(xml_name, members)

    def test_the_log_banner_comes_after_the_zip_lines(self):
        """`_create_zip_file` appends its own lines, so it must run first."""
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            processor = SurveyGenProcessor(config_for(workbook_file(tmp), out))
            run_quietly(processor)
            log = processor.logstring
            self.assertLess(
                max(i for i, line in enumerate(log) if "Added to zip" in line),
                next(i for i, line in enumerate(log) if line == "End of log file"),
            )


class FailedValidationTests(unittest.TestCase):
    def test_invalid_xml_leaves_no_manifest_and_no_loose_files(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            processor = SurveyGenProcessor(config_for(workbook_file(tmp), out))
            # Stand in for XML that ET rejects. Escaping now makes malformed
            # output very hard to produce on purpose, which is the point -- but
            # the gate still has to hold if anything ever slips through.
            processor._validate_xml_syntax = lambda path: False

            self.assertEqual(1, run_quietly(processor))

            names = sorted(p.name for p in out.iterdir())
            self.assertEqual(["gistlogfile.txt"], names)
            self.assertFalse((out / "survey_manifest.gistx").exists())
            self.assertFalse((out / "test_survey.zip").exists())

    def test_the_log_says_the_output_was_discarded(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            processor = SurveyGenProcessor(config_for(workbook_file(tmp), out))
            processor._validate_xml_syntax = lambda path: False
            run_quietly(processor)
            self.assertTrue(
                any("Discarded incomplete output" in line for line in processor.logstring),
                "\n".join(processor.logstring),
            )


class CrfsErrorTests(unittest.TestCase):
    def test_a_bad_crfs_header_stops_the_build(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            headers = list(CRFS_COLUMN_NAMES)
            headers.insert(4, "notes")
            excel = workbook_file(tmp, crfs_headers=headers)

            processor = SurveyGenProcessor(config_for(excel, out))
            self.assertEqual(1, run_quietly(processor))
            self.assertEqual(["gistlogfile.txt"], [p.name for p in out.iterdir()])
            self.assertTrue(
                any("header names" in line for line in processor.logstring),
                "\n".join(processor.logstring),
            )

    def test_a_bad_idconfig_stops_the_build(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            rows = [[
                10, "enrollee", "Enrollee", "subjid", "{bad json", 1,
                "", "", "", "", "", "", "", "", "",
            ]]
            excel = workbook_file(tmp, crfs_rows=rows)

            processor = SurveyGenProcessor(config_for(excel, out))
            self.assertEqual(1, run_quietly(processor))
            self.assertFalse((out / "test_survey.zip").exists())


if __name__ == "__main__":
    unittest.main()


def workbook_without_crfs(directory):
    """A dictionary whose `crfs` sheet is missing entirely."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "enrollee_dd"
    worksheet.append(HEADERS)
    worksheet.append(numeric_row("age"))
    path = Path(directory) / "no_crfs.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def workbook_with_sheets(directory, sheet_tables, crfs_rows, extra_rows=None):
    """A dictionary whose `_dd` sheets and `crfs` rows can be set apart."""
    workbook = Workbook()
    first = True
    for table in sheet_tables:
        worksheet = workbook.active if first else workbook.create_sheet()
        worksheet.title = f"{table}_dd"
        worksheet.append(HEADERS)
        # `extra_rows` defaults to a `subjid` row because the default crfs
        # rows name it as their primarykey, which now has to be a real column
        # on the form. A caller whose crfs rows name something else passes its
        # own rows -- an orphaned automatic `subjid` would otherwise trip the
        # separate "automatic field has no calculation" check, since only
        # fields the crfs row supplies are exempt from it.
        for extra in (
            extra_rows
            if extra_rows is not None
            else [row("subjid", "automatic", "text", "Subject ID")]
        ):
            worksheet.append(extra)
        worksheet.append(numeric_row("age"))
        first = False

    crfs = workbook.create_sheet("crfs")
    crfs.append(list(CRFS_COLUMN_NAMES))
    for r in crfs_rows:
        crfs.append(r)

    path = Path(directory) / "cross_check.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def crf_row(tablename, parenttable="", isbase=1):
    return [10, tablename, tablename.title(), "subjid", "", isbase, "",
            parenttable, "", "", "", "", "", "", ""]


def log_of(processor):
    return "\n".join(processor.logstring)


class ConfigValidationTests(unittest.TestCase):
    """A config key with no value produced a different unhelpful failure each
    time, and config.json is gitignored -- so a fresh clone running the
    documented `python main.py` hit one of them rather than a message naming
    the file to copy."""

    def build(self, tmp, **overrides):
        out = Path(tmp) / "out"
        config = config_for(workbook_file(tmp), out)
        for key, value in overrides.items():
            setattr(config, key, value)
        return SurveyGenProcessor(config), out

    def test_a_blank_excel_file_is_named_rather_than_crashing(self):
        # `Path("")` is `Path(".")`, whose `.exists()` is True, so the
        # missing-file check waved it through and `load_workbook(".")` then
        # raised an unhandled exception.
        with TemporaryDirectory() as tmp:
            processor, _ = self.build(tmp, excelFile="")

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("'excelFile' is missing or blank", log_of(processor))

    def test_a_blank_survey_id_is_refused(self):
        # It used to produce a file called ".zip".
        with TemporaryDirectory() as tmp:
            processor, _ = self.build(tmp, surveyId="")

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("'surveyId' is missing or blank", log_of(processor))

    def test_a_blank_database_name_is_refused(self):
        with TemporaryDirectory() as tmp:
            processor, _ = self.build(tmp, databaseName="")

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("'databaseName' is missing or blank", log_of(processor))

    def test_a_database_name_without_the_sqlite_suffix_is_refused(self):
        with TemporaryDirectory() as tmp:
            processor, _ = self.build(tmp, databaseName="test_database")

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("does not end in '.sqlite'", log_of(processor))

    def test_the_log_is_still_written_when_the_output_directory_does_not_exist(self):
        # The config-error path runs before `run` creates outputPath.
        with TemporaryDirectory() as tmp:
            processor, out = self.build(tmp, surveyName="")

            run_quietly(processor)
            self.assertTrue((out / "gistlogfile.txt").exists())

    def test_a_complete_config_is_untouched(self):
        with TemporaryDirectory() as tmp:
            processor, _ = self.build(tmp)

            self.assertEqual(0, run_quietly(processor), log_of(processor))


class DatabaseNameStabilityTests(unittest.TestCase):
    """databaseName must not move between survey versions: the app opens one
    database per databaseName, and a new one restarts the subject-ID counter.
    """

    def test_naming_the_database_after_the_survey_version_warns(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            config = config_for(workbook_file(tmp), out)
            config.databaseName = f"{config.surveyId}.sqlite"
            processor = SurveyGenProcessor(config)

            self.assertEqual(0, run_quietly(processor))
            self.assertIn("will change every time the survey is revised", log_of(processor))

    def test_a_database_named_after_the_study_does_not_warn(self):
        # What both live configs do: the database keeps the original date
        # while the surveyId moves on. That must stay silent.
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            config = config_for(workbook_file(tmp), out)
            config.surveyId = "avert_bf_2026_08_31"
            config.databaseName = "avert_bf_2026_07-13.sqlite"
            processor = SurveyGenProcessor(config)

            self.assertEqual(0, run_quietly(processor))
            self.assertNotIn("WARNING - databaseName", log_of(processor))


class CrfsCrossCheckTests(unittest.TestCase):
    """The crfs rows and the `_dd` worksheets are two halves of one statement
    and were never compared, so every mismatch was silent."""

    def test_a_missing_crfs_worksheet_is_an_error(self):
        # It used to yield `crfs: []` -- a survey with zero forms, reported as
        # SUCCESS, and unusable on a device that lists forms from that table.
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            processor = SurveyGenProcessor(config_for(workbook_without_crfs(tmp), out))

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("has no 'crfs' worksheet", log_of(processor))

    def test_a_tablename_with_no_worksheet_is_an_error(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            excel = workbook_with_sheets(tmp, ["enrollee"], [crf_row("enrolee")])
            processor = SurveyGenProcessor(config_for(excel, out))

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("no 'enrolee_dd' or 'enrolee_xml' worksheet", log_of(processor))

    def test_a_worksheet_with_no_crfs_row_is_an_error(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            excel = workbook_with_sheets(tmp, ["enrollee", "visit"], [crf_row("enrollee")])
            processor = SurveyGenProcessor(config_for(excel, out))

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("defines a form named 'visit', but no", log_of(processor))

    def test_a_parenttable_naming_no_declared_form_is_an_error(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            excel = workbook_with_sheets(
                tmp, ["enrollee", "visit"],
                [crf_row("enrollee"), crf_row("visit", parenttable="enrollees", isbase=0)],
            )
            processor = SurveyGenProcessor(config_for(excel, out))

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("names 'enrollees' as its parenttable", log_of(processor))

    def test_a_matching_pair_of_forms_is_accepted(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            excel = workbook_with_sheets(
                tmp, ["enrollee", "visit"],
                [crf_row("enrollee"), crf_row("visit", parenttable="enrollee", isbase=0)],
            )
            processor = SurveyGenProcessor(config_for(excel, out))

            self.assertEqual(0, run_quietly(processor), log_of(processor))


def csv_workbook(directory, responses, table="enrollee"):
    """One question whose responses come from somewhere outside the workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{table}_dd"
    worksheet.append(HEADERS)
    # primarykey has to name a real column on the form, and `crf_row` names
    # `subjid`, so the sheet declares it alongside the question under test.
    worksheet.append(row("subjid", "automatic", "text", "Subject ID"))
    worksheet.append(
        ["village", "combobox", "integer", "Which village?", "", responses,
         "", "", "", "", "", "", "", ""]
    )

    crfs = workbook.create_sheet("crfs")
    crfs.append(list(CRFS_COLUMN_NAMES))
    crfs.append(crf_row(table))

    path = Path(directory) / "csv_dictionary.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def csv_dir_with(directory, names):
    csv_dir = Path(directory) / "csvs"
    csv_dir.mkdir(exist_ok=True)
    for name in names:
        (csv_dir / name).write_text("code,name\n1,Kirembe\n")
    return csv_dir


def packaged_csvs(out, survey_id="test_survey"):
    from zipfile import ZipFile

    with ZipFile(out / f"{survey_id}.zip") as archive:
        return sorted(n for n in archive.namelist() if n.endswith(".csv"))


class CsvPackagingTests(unittest.TestCase):
    """`_create_zip_file` used to glob every *.csv in `config.csvFiles`.

    The live configs point that at a Dropbox study folder, so an unrelated
    file sitting there -- an export of already-collected records, say -- was
    bundled into the package and deployed to every tablet.
    """

    def build(self, tmp, responses, csv_names):
        out = Path(tmp) / "out"
        config = config_for(csv_workbook(tmp, responses), out)
        config.csvFiles = str(csv_dir_with(tmp, csv_names))
        return SurveyGenProcessor(config), out

    def test_an_unreferenced_csv_is_left_out_of_the_package(self):
        with TemporaryDirectory() as tmp:
            processor, out = self.build(
                tmp,
                "source:csv\nfile:villages.csv\ndisplay:name\nvalue:code",
                ["villages.csv", "hh_members_export_2026_08_30.csv"],
            )

            self.assertEqual(0, run_quietly(processor), log_of(processor))
            self.assertEqual(["villages.csv"], packaged_csvs(out))

    def test_what_was_left_out_is_named_in_the_log(self):
        # A reference this cannot see must be visible rather than silent.
        with TemporaryDirectory() as tmp:
            processor, _ = self.build(
                tmp,
                "source:csv\nfile:villages.csv\ndisplay:name\nvalue:code",
                ["villages.csv", "stray.csv"],
            )
            run_quietly(processor)

            self.assertIn("Skipped (unreferenced): stray.csv", log_of(processor))

    def test_a_referenced_file_that_is_absent_is_an_error(self):
        # It used to ship a package whose combobox is simply empty on the
        # device, with nothing on screen to explain it.
        with TemporaryDirectory() as tmp:
            processor, _ = self.build(
                tmp,
                "source:csv\nfile:villagess.csv\ndisplay:name\nvalue:code",
                ["villages.csv"],
            )

            self.assertEqual(1, run_quietly(processor))
            self.assertIn("sources its responses from 'villagess.csv'", log_of(processor))

    def test_a_csv_backed_table_named_only_by_a_database_source_is_kept(self):
        # DbService creates one table per imported CSV, so `table:villages`
        # is a reference to villages.csv even with no `file:` line. Dropping
        # it would break the survey -- the failure this selection must not
        # cause.
        with TemporaryDirectory() as tmp:
            processor, out = self.build(
                tmp,
                "source:database\ntable:villages\ndisplay:name\nvalue:code",
                ["villages.csv", "stray.csv"],
            )

            self.assertEqual(0, run_quietly(processor), log_of(processor))
            self.assertEqual(["villages.csv"], packaged_csvs(out))

    def test_a_csv_backed_table_named_only_in_calculation_sql_is_kept(self):
        # Both live dictionaries do exactly this: villages.csv is read by a
        # `<calculation type='query'>` that says `FROM villages`.
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "enrollee_dd"
            worksheet.append(HEADERS)
            worksheet.append(
                ["mrcname", "automatic", "text", "", "", 
                 "calc:query\nsql:SELECT distinct mrcname FROM villages WHERE mrccode = @mrccode\n"
                 "parameter:@mrccode=mrccode",
                 "", "", "", "", "", "", "", ""]
            )
            worksheet.append(numeric_row("mrccode"))
            # crf_row names `subjid` as the primarykey.
            worksheet.append(row("subjid", "automatic", "text", "Subject ID"))
            crfs = workbook.create_sheet("crfs")
            crfs.append(list(CRFS_COLUMN_NAMES))
            crfs.append(crf_row("enrollee"))
            excel = Path(tmp) / "sql_dictionary.xlsx"
            workbook.save(excel)
            workbook.close()

            config = config_for(excel, out)
            config.csvFiles = str(csv_dir_with(tmp, ["villages.csv", "stray.csv"]))
            processor = SurveyGenProcessor(config)

            self.assertEqual(0, run_quietly(processor), log_of(processor))
            self.assertEqual(["villages.csv"], packaged_csvs(out))

    def test_no_csv_reference_at_all_packages_nothing(self):
        with TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            config = config_for(workbook_file(tmp), out)
            config.csvFiles = str(csv_dir_with(tmp, ["stray.csv"]))
            processor = SurveyGenProcessor(config)

            self.assertEqual(0, run_quietly(processor), log_of(processor))
            self.assertEqual([], packaged_csvs(out))


class CrfsFieldReferenceTests(unittest.TestCase):
    """Every field name a crfs row mentions must name a real column.

    `crf_reader` validates the header row and the idconfig JSON and nothing
    else, so a typo in any of these cells used to produce a clean build and a
    manifest that was valid and wrong. Worse, these same names are consumed as
    an exemption list by `_supplied_auto_fields`, so a misspelled `primarykey`
    both went unreported *and* suppressed the "automatic field has no
    calculation" error for the misspelling while the real field still errored.
    """

    def build(self, tmp, sheets, crfs_rows, extra_rows=None):
        out = Path(tmp) / "out"
        path = workbook_with_sheets(tmp, sheets, crfs_rows, extra_rows)
        processor = SurveyGenProcessor(config_for(path, out))
        run_quietly(processor)
        return processor

    def test_a_primarykey_naming_no_real_column_is_an_error(self):
        with TemporaryDirectory() as tmp:
            processor = self.build(
                tmp,
                ["enrollee"],
                [[10, "enrollee", "Enrollee", "nosuchfield", "", 1, "", "", "",
                  "", "", "", "", "", ""]],
            )

            self.assertTrue(processor.errorsEncountered)
            self.assertIn("names 'nosuchfield' as its primarykey", log_of(processor))

    def test_an_incrementfield_naming_no_real_column_is_an_error(self):
        with TemporaryDirectory() as tmp:
            processor = self.build(
                tmp,
                ["enrollee", "visit"],
                [
                    [10, "enrollee", "Enrollee", "subjid", "", 1, "", "", "",
                     "", "", "", "", "", ""],
                    [20, "visit", "Visit", "subjid", "", 0, "subjid",
                     "enrollee", "nosuchfield", "", "", "", "", "", ""],
                ],
            )

            self.assertTrue(processor.errorsEncountered)
            self.assertIn(
                "names 'nosuchfield' as its incrementfield", log_of(processor)
            )

    def test_a_linkingfield_missing_from_the_parent_is_an_error(self):
        # The foreign key's actual precondition, and the one rule neither this
        # tool nor the portal checked. The app creates each child table with
        # FOREIGN KEY (linkingfield) REFERENCES parent(linkingfield), so the
        # column has to exist on both sides -- and both implementations only
        # ever looked at the child's own fields.
        with TemporaryDirectory() as tmp:
            processor = self.build(
                tmp,
                ["enrollee", "visit"],
                [
                    [10, "enrollee", "Enrollee", "subjid", "", 1, "", "", "",
                     "", "", "", "", "", ""],
                    # `age` exists on both sheets, so the child-side check
                    # passes; make the parent-side one fail instead.
                    [20, "visit", "Visit", "subjid", "", 0, "visitcode",
                     "enrollee", "", "", "", "", "", "", ""],
                ],
            )

            self.assertTrue(processor.errorsEncountered)
            self.assertIn("names 'visitcode' as its linkingfield", log_of(processor))

    def test_a_repeat_count_field_missing_from_the_parent_is_an_error(self):
        with TemporaryDirectory() as tmp:
            processor = self.build(
                tmp,
                ["enrollee", "visit"],
                [
                    [10, "enrollee", "Enrollee", "subjid", "", 1, "", "", "",
                     "", "", "", "", "", ""],
                    [20, "visit", "Visit", "subjid", "", 0, "subjid",
                     "enrollee", "", "", "nvisits", "", "", "", ""],
                ],
            )

            self.assertTrue(processor.errorsEncountered)
            self.assertIn(
                "names 'nvisits' as its repeat_count_field", log_of(processor)
            )

    def test_an_entry_condition_on_an_unknown_parent_field_is_an_error(self):
        with TemporaryDirectory() as tmp:
            processor = self.build(
                tmp,
                ["enrollee", "visit"],
                [
                    [10, "enrollee", "Enrollee", "subjid", "", 1, "", "", "",
                     "", "", "", "", "", ""],
                    [20, "visit", "Visit", "subjid", "", 0, "subjid",
                     "enrollee", "", "", "", "", "", "", "enrolled=1"],
                ],
            )

            self.assertTrue(processor.errorsEncountered)
            self.assertIn("entry_condition on 'enrolled'", log_of(processor))

    def test_a_system_variable_is_a_known_field(self):
        # The generator injects these at write time, so they are absent from
        # the sheet but present in the table.
        with TemporaryDirectory() as tmp:
            processor = self.build(
                tmp,
                ["enrollee"],
                [[10, "enrollee", "Enrollee", "uniqueid", "", 1, "", "", "",
                  "", "", "", "", "", ""]],
                extra_rows=[],
            )

            self.assertFalse(processor.errorsEncountered, log_of(processor))

    def test_a_linking_field_need_not_be_the_parent_primary_key(self):
        # AVERT's real shape: vaccination_status links to enrollee on
        # `barcode` while enrollee is keyed on `subjid`. The app declares the
        # uniqueness its foreign key needs over whichever columns children
        # reference, so this must build clean.
        with TemporaryDirectory() as tmp:
            processor = self.build(
                tmp,
                ["enrollee", "visit"],
                [
                    # enrollee is keyed on `age` but linked to on `barcode`.
                    [10, "enrollee", "Enrollee", "age", "", 1, "barcode", "",
                     "", "", "", "", "", "", ""],
                    [20, "visit", "Visit", "barcode", "", 0, "barcode",
                     "enrollee", "", "", "", "", "", "", ""],
                ],
                extra_rows=[row("barcode", "text", "text", "Barcode", maxchars="13")],
            )

            self.assertFalse(processor.errorsEncountered, log_of(processor))

    def test_a_primarykey_that_does_not_match_the_counter_only_warns(self):
        # The child counter groups siblings by linkingfield, so a primarykey
        # built from anything else describes a different grouping than the
        # database enforces. Worth flagging -- but a real dictionary already
        # surprised us on this shape, so it does not block a build.
        with TemporaryDirectory() as tmp:
            processor = self.build(
                tmp,
                ["enrollee", "visit"],
                [
                    [10, "enrollee", "Enrollee", "subjid", "", 1, "", "", "",
                     "", "", "", "", "", ""],
                    [20, "visit", "Visit", "subjid", "", 0, "subjid",
                     "enrollee", "age", "", "", "", "", "", ""],
                ],
            )

            self.assertFalse(processor.errorsEncountered, log_of(processor))
            self.assertIn("WARNING - crfs: Form 'visit'", log_of(processor))
            self.assertIn("subjid,age", log_of(processor))
