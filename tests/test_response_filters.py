import json
from pathlib import Path
import subprocess
import sys
from tempfile import TemporaryDirectory
import unittest
import xml.etree.ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook

from crf_reader import CRFS_COLUMN_NAMES
from excel_reader import ExcelReader


HEADERS = [
    "FieldName",
    "QuestionType",
    "FieldType",
    "QuestionText",
    "MaxCharacters",
    "Responses",
    "LowerRange",
    "UpperRange",
    "LogicCheck",
    "DontKnow",
    "Refuse",
    "NA",
    "Skip",
    "Comments",
]


def _question_row(fieldname, responses, question_type="radio", field_type="integer"):
    return [
        fieldname,
        question_type,
        field_type,
        f"Question {fieldname}",
        "",
        responses,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]


class ResponseFilterOperatorTests(unittest.TestCase):
    """The filter operator must survive the trip from Excel to XML intact.

    A filter whose operator is not recognized still matches the (optional)
    operator group, so a mistake here is silent: the operator quietly becomes
    '=' and the rest of the line is swallowed into the value.
    """

    def _filters_for(self, filter_lines):
        responses = "\n".join(
            ["source:database", "table:hh_members"]
            + filter_lines
            + ["display:participantsname", "value:linenum"]
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "nets_dd"
        worksheet.append(HEADERS)
        worksheet.append(_question_row("who", responses))

        reader = ExcelReader()
        questions = reader.create_question_list(worksheet)
        self.assertFalse(reader.errorsEncountered, "\n".join(reader.logstring))
        return questions[0].responseFilters

    def test_not_in_is_kept_as_an_operator(self):
        filters = self._filters_for(["filter:linenum not in [[used_linenums]]"])

        self.assertEqual(len(filters), 1)
        self.assertEqual(filters[0].column, "linenum")
        self.assertEqual(filters[0].operator, "not in")
        self.assertEqual(filters[0].value, "[[used_linenums]]")

    def test_in_is_kept_as_an_operator(self):
        filters = self._filters_for(["filter:linenum in [[used_linenums]]"])

        self.assertEqual(filters[0].operator, "in")
        self.assertEqual(filters[0].value, "[[used_linenums]]")

    def test_irregular_spacing_and_case_are_normalized(self):
        filters = self._filters_for(["filter:linenum NOT  IN [[used_linenums]]"])

        self.assertEqual(filters[0].operator, "not in")
        self.assertEqual(filters[0].value, "[[used_linenums]]")

    def test_greater_or_equal_is_not_split_into_greater_than(self):
        filters = self._filters_for(["filter:census_age >= 15"])

        self.assertEqual(filters[0].operator, "&gt;=")
        self.assertEqual(filters[0].value, "15")

    def test_less_or_equal_is_not_split_into_less_than(self):
        filters = self._filters_for(["filter:census_age <= 15"])

        self.assertEqual(filters[0].operator, "&lt;=")
        self.assertEqual(filters[0].value, "15")

    def test_not_equal_is_escaped_for_xml(self):
        # '<' is not legal inside an XML attribute value.
        filters = self._filters_for(["filter:census_age <> 15"])

        self.assertEqual(filters[0].operator, "&lt;&gt;")
        self.assertEqual(filters[0].value, "15")

    def test_simple_operators_are_unchanged(self):
        for line, operator, value in [
            ("filter:hhid = [[hhid]]", "=", "[[hhid]]"),
            ("filter:sex != 1", "!=", "1"),
            ("filter:age > 5", "&gt;", "5"),
            ("filter:age < 5", "&lt;", "5"),
        ]:
            with self.subTest(line=line):
                filters = self._filters_for([line])
                self.assertEqual(filters[0].operator, operator)
                self.assertEqual(filters[0].value, value)

    def test_missing_operator_still_defaults_to_equals(self):
        filters = self._filters_for(["filter:hhid [[hhid]]"])

        self.assertEqual(filters[0].operator, "=")
        self.assertEqual(filters[0].value, "[[hhid]]")

    def test_multiple_filters_are_kept_in_order(self):
        filters = self._filters_for(
            [
                "filter:hhid = [[hhid]]",
                "filter:linenum not in [[used_linenums]]",
            ]
        )

        self.assertEqual([f.operator for f in filters], ["=", "not in"])


class ResponseFilterXmlTests(unittest.TestCase):
    """End-to-end: the operator must reach the generated XML unchanged."""

    def test_not_in_filter_is_written_to_the_xml(self):
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            workbook_path = temp_path / "filters.xlsx"
            output_path = temp_path / "output"
            config_path = temp_path / "config.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "nets_dd"
            worksheet.append(HEADERS)

            # A crfs sheet, because the processor now refuses a workbook
            # without one -- the manifest would otherwise declare a survey
            # with no forms.
            crfs = workbook.create_sheet("crfs")
            crfs.append(list(CRFS_COLUMN_NAMES))
            crfs.append([10, "nets", "Nets", "netid", "", 1, "", "", "", "", "", "", "", "", ""])
            worksheet.append(
                _question_row(
                    "used_linenums",
                    "calc:query\n"
                    "sql:SELECT group_concat(sleptunder) FROM nets WHERE hhid = @hhid "
                    "AND netnum != @netnum AND sleptunder IS NOT NULL\n"
                    "param:@hhid = hhid\n"
                    "param:@netnum = netnum",
                    question_type="automatic",
                    field_type="text",
                )
            )
            worksheet.append(
                _question_row(
                    "sleptunder",
                    "source:database\n"
                    "table:hh_members\n"
                    "filter:hhid = [[hhid]]\n"
                    "filter:linenum not in [[used_linenums]]\n"
                    "display:participantsname\n"
                    "value:linenum",
                    question_type="checkbox",
                    field_type="text",
                )
            )
            workbook.save(workbook_path)

            config_path.write_text(
                json.dumps(
                    {
                        "excelFile": str(workbook_path),
                        "csvFiles": "",
                        "outputPath": str(output_path),
                        "surveyName": "Response Filter Test",
                        "surveyId": "response_filter_test",
                        "databaseName": "response_filter_test_data.sqlite",
                    }
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [sys.executable, "main.py", "--config", str(config_path)],
                cwd=Path(__file__).resolve().parents[1],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            log_text = (output_path / "gistlogfile.txt").read_text(encoding="utf-8")
            self.assertNotIn("ERROR", log_text)

            with ZipFile(output_path / "response_filter_test.zip") as archive:
                # Parsing at all proves the SQL and operators are XML-safe.
                root = ET.fromstring(archive.read("nets.xml"))

            filters = root.findall(
                ".//question[@fieldname='sleptunder']/responses/filter"
            )
            self.assertEqual(len(filters), 2)
            self.assertEqual(filters[1].get("column"), "linenum")
            self.assertEqual(filters[1].get("operator"), "not in")
            self.assertEqual(filters[1].get("value"), "[[used_linenums]]")

            sql = root.find(
                ".//question[@fieldname='used_linenums']/calculation/sql"
            )
            self.assertIn("group_concat(sleptunder)", sql.text)
            self.assertIn("netnum != @netnum", sql.text)


if __name__ == "__main__":
    unittest.main()
