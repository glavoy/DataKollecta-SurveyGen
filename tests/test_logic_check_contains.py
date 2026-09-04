import json
from pathlib import Path
import subprocess
import sys
from tempfile import TemporaryDirectory
import unittest
import xml.etree.ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook

from crf_reader import CRFS_COLUMN_NAMES
from excel_reader import ExcelReader


HEADERS = [
    "FieldName",
    "QuestionType",
    "FieldType",
    "QuestionText",
    "MaxCharacters",
    "Responses",
    "LowerRange",
    "UpperRange",
    "LogicCheck",
    "DontKnow",
    "Refuse",
    "NA",
    "Skip",
    "Comments",
]


class LogicCheckContainsTests(unittest.TestCase):
    LOGIC_CHECK = (
        "symptoms contains '99' and (symptoms contains '1' or symptoms contains '2' or "
        "symptoms contains '3' or symptoms contains '4' or symptoms contains '5' or "
        "symptoms contains '6' or symptoms contains '7' or symptoms contains '8' or "
        "symptoms contains '9' or symptoms contains '10' or symptoms contains '11' or "
        "symptoms contains '12'); 'Cannot select \"Child does not have any symptoms\" "
        "together with other symptoms'"
    )

    @staticmethod
    def _add_questions(worksheet):
        worksheet.append(HEADERS)
        worksheet.append(
            [
                "symptoms",
                "checkbox",
                "text",
                "Select all symptoms",
                "",
                "1:Fever\n2:Cough\n3:Symptom 3\n4:Symptom 4\n5:Symptom 5\n6:Symptom 6\n"
                "7:Symptom 7\n8:Symptom 8\n9:Symptom 9\n10:Symptom 10\n11:Symptom 11\n"
                "12:Symptom 12\n99:Child does not have any symptoms",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        worksheet.append(
            [
                "symptoms_check",
                "information",
                "n/a",
                "Check symptom selections",
                "",
                "",
                "",
                "",
                LogicCheckContainsTests.LOGIC_CHECK,
                "",
                "",
                "",
                "",
                "",
            ]
        )

    def test_contains_and_does_not_contain_are_not_treated_as_field_names(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "symptoms_dd"
        self._add_questions(worksheet)

        reader = ExcelReader()
        reader.create_question_list(worksheet)

        self.assertFalse(reader.errorsEncountered, "\n".join(reader.logstring))

    def test_processor_generates_logic_check_xml_for_contains_expression(self):
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            workbook_path = temp_path / "contains_logic_check.xlsx"
            output_path = temp_path / "output"
            config_path = temp_path / "config.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "symptoms_dd"
            self._add_questions(worksheet)

            # A crfs sheet, because the processor now refuses a workbook
            # without one -- the manifest would otherwise declare a survey
            # with no forms.
            crfs = workbook.create_sheet("crfs")
            crfs.append(list(CRFS_COLUMN_NAMES))
            # primarykey has to name a real column on the form now, and
            # `symptoms` is what this sheet actually declares.
            crfs.append([10, "symptoms", "Symptoms", "symptoms", "", 1, "", "", "", "", "", "", "", "", ""])

            workbook.save(workbook_path)

            config_path.write_text(
                json.dumps(
                    {
                        "excelFile": str(workbook_path),
                        "csvFiles": "",
                        "outputPath": str(output_path),
                        "surveyName": "Contains LogicCheck Test",
                        "surveyId": "contains_logic_check_test",
                        "databaseName": "contains_logic_check_test_data.sqlite",
                    }
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [sys.executable, "main.py", "--config", str(config_path)],
                cwd=Path(__file__).resolve().parents[1],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            log_text = (output_path / "gistlogfile.txt").read_text(encoding="utf-8")
            self.assertNotIn("ERROR", log_text)

            with ZipFile(output_path / "contains_logic_check_test.zip") as archive:
                root = ET.fromstring(archive.read("symptoms.xml"))

            logic_check = root.find(".//question[@fieldname='symptoms_check']/logic_check")
            self.assertIsNotNone(logic_check)
            rendered_expression = " ".join((logic_check.text or "").split())
            self.assertIn("symptoms contains '99'", rendered_expression)
            self.assertIn("symptoms contains '12'", rendered_expression)
            self.assertIn('Cannot select "Child does not have any symptoms" together with other symptoms', rendered_expression)


if __name__ == "__main__":
    unittest.main()
