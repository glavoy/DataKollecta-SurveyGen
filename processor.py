from __future__ import annotations

import json
import re
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from crf_reader import CrfReader
from excel_reader import ExcelReader
from json_generator import JsonGenerator
from models import AppConfig, Question, ResponseSourceType, SurveyManifest
from xml_generator import XmlGenerator


class SurveyGenProcessor:
    def __init__(self, config: AppConfig):
        self.config = config
        self.errorsEncountered = False
        self.logstring: list[str] = []
        self.generated_files: list[Path] = []
        self.question_list_cache: dict[str, list[Question]] = {}

    # Config keys that have no sensible default. Each one produced a
    # different unhelpful failure when left out of config.json, and because
    # config.json is gitignored, a fresh clone running the documented
    # `python main.py` hit one of them rather than "copy config.sample.json".
    REQUIRED_CONFIG_KEYS = ("excelFile", "outputPath", "surveyId", "surveyName", "databaseName")

    def _validate_config(self) -> list[str]:
        """Config problems, as error lines. Empty means the config is usable."""
        errors: list[str] = []

        for key in self.REQUIRED_CONFIG_KEYS:
            if not str(getattr(self.config, key, "") or "").strip():
                errors.append(
                    f"ERROR - Config: '{key}' is missing or blank in the config file. "
                    "Copy config.sample.json and fill in every value."
                )

        # Checked separately from "is it blank", because a blank excelFile is
        # not merely unset -- `Path("")` is `Path(".")`, whose `.exists()` is
        # True, so the missing-file check below waved it through and
        # `load_workbook(".")` then raised an unhandled exception.
        database_name = str(self.config.databaseName or "").strip()
        if database_name and not database_name.endswith(".sqlite"):
            errors.append(
                f"ERROR - Config: 'databaseName' is '{database_name}', which does not end in "
                "'.sqlite'. The app opens one database file per databaseName and expects that "
                "suffix."
            )

        return errors

    def _check_database_name_stability(self) -> None:
        """Warn when databaseName looks like it was named after this version.

        The platform README calls a stable databaseName the one rule that
        cannot be broken: the subject-ID counter is derived from MAX(...) over
        the survey's own table, so a new databaseName silently restarts it and
        produces duplicate subject IDs. The mistake that causes it is naming
        the database after the survey rather than after the study, because the
        surveyId carries the version date and therefore moves every revision.

        Deliberately NOT attempted here: detecting that databaseName changed
        since the last build. That needs state across runs which this tool
        does not keep, and guessing at it from the output directory would be
        worse than saying nothing. Both live configs name a database after the
        study and hold the original date -- exactly right -- so a blanket
        "contains a date" warning would fire on correct configs.
        """
        database_name = str(self.config.databaseName or "").strip()
        survey_id = str(self.config.surveyId or "").strip()
        if database_name and survey_id and database_name == f"{survey_id}.sqlite":
            self.logstring.append(
                f"WARNING - databaseName: '{database_name}' is the surveyId with '.sqlite' on the "
                "end, so it will change every time the survey is revised. The app opens one "
                "database per databaseName, and a new one restarts the subject-ID counter -- name "
                "it after the study, not after this version, and never change it again."
            )

    def run(self) -> int:
        self.logstring = [f"Log file for: {self.config.excelFile}"]

        config_errors = self._validate_config()
        if config_errors:
            self.logstring.extend(config_errors)
            self.logstring.extend(self._end_of_log_banner())
            self._write_logfile()
            print("ERRORS FOUND: The config file is incomplete.")
            for line in config_errors:
                print(f"  {line}")
            return 1

        self._check_database_name_stability()

        excel_path = Path(self.config.excelFile)
        output_path = Path(self.config.outputPath)
        output_path.mkdir(parents=True, exist_ok=True)

        if not excel_path.exists():
            self.logstring.append("ERROR: Excel file not found!")
            self.logstring.append(str(excel_path))
            self.logstring.extend(self._end_of_log_banner())
            self._write_logfile()
            print(f"ERROR: Excel file not found: {excel_path}")
            print("Please check the 'excelFile' path in config.json.")
            return 1

        workbook = load_workbook(filename=excel_path, data_only=False)

        try:
            worksheets = [ws for ws in workbook.worksheets if ws.title.endswith("_dd") or ws.title.endswith("_xml")]

            # Read the crfs sheet first: it tells us which automatic fields the
            # manifest fills in (linking field, increment field, primary key,
            # idconfig parts), so they are not reported as missing a calculation.
            crfs_ws = workbook["crfs"] if "crfs" in workbook.sheetnames else None
            if crfs_ws is not None:
                crfs, crf_errors = CrfReader.read_crfs_worksheet(crfs_ws)
                if crf_errors:
                    self.logstring.append("\rChecking worksheet: 'crfs'")
                    self.logstring.extend(crf_errors)
                    self.errorsEncountered = True
            else:
                # Used to be a silent `crfs = []`, which produced a manifest
                # declaring a survey with zero forms -- reported as SUCCESS,
                # and unusable on the device, which drives its whole screen
                # from the crfs table.
                crfs = []
                self.logstring.append("\rChecking worksheet: 'crfs'")
                self.logstring.append(
                    "ERROR - crfs: The workbook has no 'crfs' worksheet. It is what tells the app "
                    "which forms exist, how they link to each other and how IDs are built; without "
                    "it the package declares a survey with no forms at all."
                )
                self.errorsEncountered = True

            self._check_crfs_against_worksheets(crfs, worksheets)
            supplied_by_table = self._supplied_auto_fields(crfs)

            for ws in worksheets:
                table = ws.title.replace("_dd", "").replace("_xml", "")
                reader = ExcelReader(supplied_auto_fields=supplied_by_table.get(table, set()))
                qlist = reader.create_question_list(ws)
                if reader.errorsEncountered:
                    self.errorsEncountered = True
                self.logstring.extend(reader.logstring)
                self.question_list_cache[ws.title] = qlist

            self._check_csv_references()

            xml_files: list[str] = []

            if not self.errorsEncountered:
                for ws_name, qlist in self.question_list_cache.items():
                    xml_generator = XmlGenerator()
                    xml_path = xml_generator.write_xml(ws_name, qlist, output_path)
                    self.logstring.extend(xml_generator.logstring)
                    self.generated_files.append(xml_path)

                    # Taken from the file that was actually written rather than
                    # re-derived from the worksheet name. The two derivations
                    # used to disagree (this one did
                    # `ws_name.replace("_dd", ".xml").replace("_xml", ".xml")`,
                    # `write_xml` sliced the suffix), so a sheet named
                    # `hh_xml_dd` put `hh.xml.xml` in the manifest while the
                    # zip held `hh_xml.xml` -- a manifest naming a file that
                    # was not in the package.
                    xml_files.append(xml_path.name)

                    if not self._validate_xml_syntax(xml_path):
                        self.errorsEncountered = True

            # Deliberately a second, separate check rather than part of the
            # block above. The XML loop can set `errorsEncountered` itself, and
            # when it did, the enclosing `if` had already been entered -- so
            # the manifest was written anyway and only the zip was skipped.
            # That left every .xml and the .gistx sitting in outputPath while
            # the console said "HAVE NOT been created", and hand-zipping those
            # files ships a package containing malformed XML.
            if not self.errorsEncountered:
                manifest = SurveyManifest(
                    surveyName=self.config.surveyName,
                    surveyId=self.config.surveyId,
                    databaseName=self.config.databaseName,
                    xmlFiles=xml_files,
                    crfs=crfs,
                )
                manifest_path = output_path / "survey_manifest.gistx"
                JsonGenerator.write_manifest(manifest_path, manifest)
                self.logstring.append("")
                self.logstring.append("Successfully generated survey_manifest.gistx")
                self.generated_files.append(manifest_path)

                self._create_zip_file()
            else:
                self._discard_generated_files()

            # After the zip, so `_create_zip_file`'s own "Added to zip" and
            # "Deleted temporary file" lines land before the banner instead of
            # after the end of the file.
            self.logstring.extend(self._end_of_log_banner())

            self._write_logfile()

            # Console equivalent of the Windows app's SUCCESS/ERRORS FOUND message boxes
            logfile_path = output_path / "gistlogfile.txt"
            if self.errorsEncountered:
                print("ERRORS FOUND: The Data Dictionary contains errors!")
                print("No package was created: the XML files and manifest have been discarded.")
                print(f"Please refer to the log file and rectify all errors: {logfile_path}")
            else:
                print("SUCCESS: Built the XML file(s) and the manifest. No errors were found.")
                print(f"All files have been packaged in: {output_path / (self.config.surveyId + '.zip')}")
                print(f"Log file: {logfile_path}")
            return 1 if self.errorsEncountered else 0
        finally:
            workbook.close()

    # Identifiers a calculation query selects from. A DB-backed response or a
    # `<calculation type='query'>` names a TABLE, and DbService creates one
    # table per imported CSV -- so `FROM villages` is a reference to
    # villages.csv even though no `file:` line mentions it. Both live
    # dictionaries do exactly this.
    SQL_TABLE_RE = re.compile(r"\b(?:from|join)\s+([A-Za-z_][A-Za-z0-9_]*)", re.IGNORECASE)

    def _referenced_csv_names(self) -> tuple[set[str], set[str]]:
        """(filenames named by `file:`, table names that may be CSV-backed).

        The two are kept apart because they carry different weight. A `file:`
        reference is unambiguous -- if the file is absent the combobox is
        empty in the field, so a missing one is an error. A table name is a
        guess: it may be a CSV-backed lookup or one of the survey's own form
        tables, and it is only ever used to *keep* a file, never to demand one.
        """
        files: set[str] = set()
        tables: set[str] = set()

        for qlist in self.question_list_cache.values():
            for q in qlist:
                if q.responseSourceType == ResponseSourceType.CSV and q.responseSourceFile:
                    files.add(q.responseSourceFile.strip())
                if q.responseSourceTable:
                    tables.add(q.responseSourceTable.strip().lower())
                for sql in self._calculation_sql(q):
                    tables.update(m.lower() for m in self.SQL_TABLE_RE.findall(sql))

        return files, tables

    @classmethod
    def _calculation_sql(cls, question) -> list[str]:
        """Every SQL string a question carries, including nested parts.

        A `case` or `math` calculation holds its queries in CalculationPart
        children rather than on the question itself, so a flat read of
        `calculationQuerySql` would miss them.
        """
        found: list[str] = []
        if getattr(question, "calculationQuerySql", ""):
            found.append(question.calculationQuerySql)

        def walk(part) -> None:
            if part is None:
                return
            if getattr(part, "querySql", ""):
                found.append(part.querySql)
            for child in getattr(part, "parts", None) or ():
                walk(child)

        for condition in getattr(question, "calculationCaseConditions", None) or ():
            walk(getattr(condition, "result", None))
            walk(getattr(condition, "value", None))
        walk(getattr(question, "calculationCaseElse", None))
        for group in ("calculationMathParts", "calculationConcatParts"):
            for child in getattr(question, group, None) or ():
                walk(child)

        return found

    def _check_crfs_against_worksheets(self, crfs: list, worksheets: list) -> None:
        """Every form must be declared once and defined once.

        The two lists sat side by side in `run` and were never compared, so
        all three mismatches were silent: a `tablename` typo declared a form
        with no XML behind it, a `_dd` sheet with no crfs row shipped XML for
        a form the app never lists, and a `parenttable` naming a table that
        does not exist broke the parent/child link at interview time.
        """
        if not crfs:
            return

        sheet_tables = {ws.title.replace("_dd", "").replace("_xml", ""): ws.title for ws in worksheets}
        crf_tables = {(crf.tablename or "").strip() for crf in crfs if (crf.tablename or "").strip()}

        for crf in crfs:
            tablename = (crf.tablename or "").strip()
            if not tablename:
                continue
            if tablename not in sheet_tables:
                self._crfs_error(
                    f"ERROR - crfs: 'tablename' is '{tablename}', but the workbook has no "
                    f"'{tablename}_dd' or '{tablename}_xml' worksheet. The manifest would declare a "
                    "form with no questionnaire behind it."
                )

            parent = (crf.parenttable or "").strip()
            if parent and parent not in crf_tables:
                self._crfs_error(
                    f"ERROR - crfs: Form '{tablename}' names '{parent}' as its parenttable, but no "
                    "crfs row declares that table. The child would have no parent to link to."
                )

        for table, sheet_title in sorted(sheet_tables.items()):
            if table not in crf_tables:
                self._crfs_error(
                    f"ERROR - crfs: Worksheet '{sheet_title}' defines a form named '{table}', but no "
                    "crfs row declares it. Its XML would be packaged and then never listed by the "
                    "app, which drives its questionnaire list from the crfs table."
                )

    def _crfs_error(self, message: str) -> None:
        """Log a crfs-level error under the same heading CrfReader uses."""
        heading = "\rChecking worksheet: 'crfs'"
        if heading not in self.logstring:
            self.logstring.append(heading)
        self.logstring.append(message)
        self.errorsEncountered = True

    @staticmethod
    def _end_of_log_banner() -> list[str]:
        return [
            "\r--------------------------------------------------------------------------------",
            "End of log file",
            "--------------------------------------------------------------------------------",
        ]

    @staticmethod
    def _supplied_auto_fields(crfs: list) -> dict[str, set[str]]:
        """Automatic fields the app fills in from the manifest, per table.

        These are populated at runtime from the crfs entry rather than by a
        calculation, so they must not be reported as missing one.
        """
        supplied: dict[str, set[str]] = {}
        for crf in crfs:
            if not crf.tablename:
                continue
            fields: set[str] = set()
            for key in (crf.linkingfield, crf.incrementfield):
                if key:
                    fields.add(key.strip())
            for key in (crf.primarykey or "").split(","):
                if key.strip():
                    fields.add(key.strip())
            if crf.idconfig and crf.idconfig.fields:
                for f in crf.idconfig.fields:
                    if f.name:
                        fields.add(f.name.strip())
            supplied[crf.tablename] = fields
        return supplied

    def _validate_xml_syntax(self, file_path: Path) -> bool:
        try:
            ET.parse(file_path)
            return True
        except ET.ParseError as ex:
            self.logstring.append(f"CRITICAL ERROR: XML Syntax Error in file '{file_path.name}'")
            self.logstring.append(f"Details: {ex}")
            return False
        except Exception as ex:
            self.logstring.append(f"CRITICAL ERROR: Could not validate XML file '{file_path.name}'")
            self.logstring.append(f"Details: {ex}")
            return False

    def _write_logfile(self) -> None:
        # Matches the C# StreamWriter output: CRLF after every line, and a
        # final WriteLine("\n") that emits "\n\r\n".
        logfile = Path(self.config.outputPath) / "gistlogfile.txt"
        # The config-error path writes a log before `run` has had a chance to
        # create outputPath, and a log nobody can read is the one thing worse
        # than the error it describes.
        logfile.parent.mkdir(parents=True, exist_ok=True)
        with logfile.open("w", encoding="utf-8", newline="") as f:
            for line in self.logstring:
                f.write(line + "\r\n")
            f.write("\n\r\n")

    def _create_zip_file(self) -> None:
        zip_file_path = Path(self.config.outputPath) / f"{self.config.surveyId}.zip"
        if zip_file_path.exists():
            zip_file_path.unlink()

        with ZipFile(zip_file_path, "w", compression=ZIP_DEFLATED) as archive:
            for file_path in self.generated_files:
                if file_path.exists():
                    archive.write(file_path, arcname=file_path.name)
                    self.logstring.append(f"Added to zip: {file_path.name}")

            for csv in self._csv_files_to_package():
                archive.write(csv, arcname=csv.name)
                self.logstring.append(f"Added to zip: {csv.name}")

        self.logstring.append("")
        self.logstring.append(f"Successfully created zip file: {zip_file_path}")

        for file_path in self.generated_files:
            if file_path.exists():
                file_path.unlink()
                self.logstring.append(f"Deleted temporary file: {file_path.name}")

    def _csv_files_to_package(self) -> list[Path]:
        """The CSVs this survey actually references, and only those.

        This used to glob every *.csv in `config.csvFiles` and package the
        lot. The local configs point that at a Dropbox study folder, so any
        unrelated CSV sitting there -- including an export of already-collected
        records -- was bundled into the zip and deployed to every tablet.

        Selection is deliberately generous, because shipping one file too many
        is a smaller failure than shipping one too few (a missing lookup means
        an empty response list in the field). A file is kept if a `file:`
        reference names it, or if its stem matches a table named by a
        `source:database` response or by a calculation query. Whatever is left
        out is listed by name in the log, so a reference this cannot see is
        visible rather than silent.
        """
        if not self.config.csvFiles:
            return []

        csv_dir = Path(self.config.csvFiles.rstrip("\\/"))
        if not (csv_dir.exists() and csv_dir.is_dir()):
            self.logstring.append(f"WARNING: CSV files directory not found: {csv_dir}")
            return []

        available = {path.name: path for path in sorted(csv_dir.glob("*.csv"))}
        referenced_files, referenced_tables = self._referenced_csv_names()

        keep: list[Path] = []
        for name, path in available.items():
            if name in referenced_files or Path(name).stem.lower() in referenced_tables:
                keep.append(path)

        if keep:
            self.logstring.append("")
            self.logstring.append("Adding CSV files to package:")

        skipped = sorted(set(available) - {p.name for p in keep})
        if skipped:
            self.logstring.append("")
            self.logstring.append(
                f"NOTE: {len(skipped)} CSV file(s) in {csv_dir} are not referenced by any "
                "question and were NOT packaged:"
            )
            for name in skipped:
                self.logstring.append(f"  Skipped (unreferenced): {name}")

        return keep

    def _check_csv_references(self) -> None:
        """A `file:` reference that names a CSV which is not there.

        Nothing checked this, so a misspelled filename shipped a package whose
        combobox is simply empty on the device -- with no error at generation
        time and nothing on screen to explain it.
        """
        referenced_files, _ = self._referenced_csv_names()
        if not referenced_files:
            return

        csv_dir = Path(self.config.csvFiles.rstrip("\\/")) if self.config.csvFiles else None
        if csv_dir is None or not (csv_dir.exists() and csv_dir.is_dir()):
            self.logstring.append(
                "ERROR - CSV: Questions reference CSV files "
                f"({', '.join(sorted(referenced_files))}), but 'csvFiles' in the config does not "
                "point at a readable directory."
            )
            self.errorsEncountered = True
            return

        available = {path.name for path in csv_dir.glob("*.csv")}
        for name in sorted(referenced_files):
            if name not in available:
                self.logstring.append(
                    f"ERROR - CSV: A question sources its responses from '{name}', but no such "
                    f"file is in {csv_dir}. The question would show an empty list in the field."
                )
                self.errorsEncountered = True

    def _discard_generated_files(self) -> None:
        """Remove part-built output after a failure.

        Cleanup used to live only inside `_create_zip_file`, which is skipped
        on failure -- so a failed run left loose .xml files (and, before the
        gate above was split in two, survey_manifest.gistx) behind for someone
        to zip by hand.
        """
        for file_path in self.generated_files:
            if file_path.exists():
                file_path.unlink()
                self.logstring.append(f"Discarded incomplete output: {file_path.name}")


def run_from_config_file(config_file: str | Path) -> int:
    cfg_path = Path(config_file)
    data = json.loads(cfg_path.read_text(encoding="utf-8"))
    config = AppConfig(
        excelFile=data.get("excelFile", ""),
        csvFiles=data.get("csvFiles", ""),
        outputPath=data.get("outputPath", ""),
        surveyName=data.get("surveyName", ""),
        surveyId=data.get("surveyId", ""),
        databaseName=data.get("databaseName", ""),
    )
    processor = SurveyGenProcessor(config)
    return processor.run()
