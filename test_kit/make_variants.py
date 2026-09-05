"""Turn one real dictionary into the twelve that sweep the repeat matrix.

    python -m test_kit.make_variants --source "/path/PRISMCOMP_CSS_...xlsx"

Each variant is the source workbook with **two cells changed** -- one child
form's `auto_start_repeat` and `repeat_enforce_count`. Nothing else moves.

That is the whole design, and it is deliberate. The alternative was authoring a
synthetic dictionary that exercises the matrix, which would have been a form
nobody ever ran: no real skips, no cascading CSV filters, no increment field
used inside a logic check, and no 149-row child. A two-cell diff from a study
that actually shipped keeps every one of those and changes only the thing under
test. It also cannot drift from reality, because reality is the input.

**Why an openpyxl round-trip is safe here.** Loading and saving a workbook does
not preserve everything Excel can store. It does not have to: the only consumer
of these files is `SurveyGenProcessor`, which reads them with openpyxl too. Any
formatting a round-trip drops was never read by the generator, so it cannot
change what the package contains. (The source workbook is opened read-only in
spirit -- it is loaded and never saved back over.)

Outputs, all gitignored:

    test_kit/variants/<prefix>_<cell>.xlsx      the workbook
    test_kit/variants/config_<cell>.json        what `main.py --config` takes
    test_kit/variants/matrix.json               matrix.py, for non-Python readers
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from crf_reader import CRFS_COLUMN_NAMES
from test_kit import matrix as matrix_module
from test_kit.matrix import MatrixCell

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUT = REPO_ROOT / "test_kit" / "variants"
DEFAULT_PREFIX = "zz_prism"

# Read by position, exactly as `CrfReader.read_crfs_worksheet` does. Deriving
# these from the shared list rather than hardcoding 12 and 13 means an inserted
# column moves both readers together instead of silently landing this script's
# writes in the wrong cells -- the failure `crf_reader`'s header check exists
# to catch.
_TABLENAME_COL = CRFS_COLUMN_NAMES.index("tablename") + 1
_AUTO_START_COL = CRFS_COLUMN_NAMES.index("auto_start_repeat") + 1
_ENFORCE_COL = CRFS_COLUMN_NAMES.index("repeat_enforce_count") + 1


class VariantError(RuntimeError):
    """The source workbook is not the shape this script can rewrite."""


@dataclass(frozen=True)
class Variant:
    cell: MatrixCell
    workbook: Path
    config: Path
    survey_id: str
    database_name: str


def _crfs_sheet(workbook):
    for name in workbook.sheetnames:
        if name.strip().lower() == "crfs":
            return workbook[name]
    raise VariantError(
        "The source workbook has no 'crfs' worksheet, so there are no repeat "
        f"columns to override. Sheets found: {', '.join(workbook.sheetnames)}"
    )


def _check_headers(sheet) -> None:
    """Refuse a sheet whose columns are not where this script will write.

    `crf_reader` reads every crfs column by position and rejects a reordered
    header outright. This script *writes* by position, so the same check has to
    happen here -- and it has to happen before anything is saved, or the
    variants are valid workbooks holding the override in the wrong column.
    """
    found = [
        (sheet.cell(row=1, column=i + 1).value or "").strip().lower()
        if isinstance(sheet.cell(row=1, column=i + 1).value, str)
        else (sheet.cell(row=1, column=i + 1).value or "")
        for i in range(len(CRFS_COLUMN_NAMES))
    ]
    if [str(f) for f in found] != list(CRFS_COLUMN_NAMES):
        raise VariantError(
            "The 'crfs' worksheet headers are not in the expected order, so "
            "writing by column position would put the override in the wrong "
            f"cell.\n  Expected: {', '.join(CRFS_COLUMN_NAMES)}\n"
            f"  Found:    {', '.join(str(f) or '(blank)' for f in found)}"
        )


def _find_child_row(sheet, child: str) -> int:
    for row_idx in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=_TABLENAME_COL).value
        if value is not None and str(value).strip() == child:
            return row_idx
    tables = [
        str(sheet.cell(row=r, column=_TABLENAME_COL).value)
        for r in range(2, sheet.max_row + 1)
        if sheet.cell(row=r, column=_TABLENAME_COL).value
    ]
    raise VariantError(
        f"No crfs row names '{child}', so there is nothing to reconfigure. "
        f"Tables in this dictionary: {', '.join(tables) or '(none)'}. "
        "Pass --child to name one of them."
    )


def build(
    source: Path,
    out_dir: Path,
    *,
    csv_dir: Path,
    output_path: Path,
    child: str = matrix_module.MATRIX_CHILD,
    prefix: str = DEFAULT_PREFIX,
    survey_name_prefix: str = "[TEST]",
) -> list[Variant]:
    """Write one workbook and one config per matrix cell. Returns what it wrote."""
    if not source.is_file():
        raise VariantError(f"Source dictionary not found: {source}")

    out_dir.mkdir(parents=True, exist_ok=True)

    variants: list[Variant] = []
    for cell in matrix_module.cells():
        # Reloaded per cell rather than deep-copied. openpyxl workbooks hold
        # live references into their own parts, so a copy is not reliably
        # independent; reloading is a few milliseconds and is obviously correct.
        workbook = load_workbook(source, data_only=False)
        try:
            sheet = _crfs_sheet(workbook)
            _check_headers(sheet)
            row_idx = _find_child_row(sheet, child)

            sheet.cell(row=row_idx, column=_AUTO_START_COL).value = cell.auto_start_repeat
            sheet.cell(row=row_idx, column=_ENFORCE_COL).value = cell.repeat_enforce_count

            survey_id = f"{prefix}_{cell.key}"
            workbook_path = out_dir / f"{survey_id}.xlsx"
            workbook.save(workbook_path)
        finally:
            workbook.close()

        database_name = f"{survey_id}.sqlite"
        config_path = out_dir / f"config_{cell.key}.json"
        config_path.write_text(
            json.dumps(
                {
                    "excelFile": str(workbook_path),
                    "csvFiles": str(csv_dir),
                    "outputPath": str(output_path),
                    # Visibly a test package wherever it is listed. The app
                    # shows surveyName in its survey picker, so anyone who
                    # side-loads one of these onto a real device sees what it
                    # is before they start typing into it.
                    "surveyName": f"{survey_name_prefix} {child} {cell.describe()}",
                    # surveyId is the zip filename and the on-device extraction
                    # folder; databaseName is the SQLite filename and the key
                    # DbService opens databases by. Distinct values are what
                    # keep twelve variants from sharing one subject-ID counter.
                    "surveyId": survey_id,
                    "databaseName": database_name,
                },
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        variants.append(
            Variant(
                cell=cell,
                workbook=workbook_path,
                config=config_path,
                survey_id=survey_id,
                database_name=database_name,
            )
        )

    (out_dir / "matrix.json").write_text(
        json.dumps(matrix_module.to_json_dict(), indent=2) + "\n", encoding="utf-8"
    )
    return variants


def _default_csv_dir(source: Path) -> Path:
    """Where the dictionary's CSVs live, if not told otherwise.

    Alongside the workbook is right for every dictionary here: PRISM's
    `villages.csv` sits next to it, and so does AVERT's.
    """
    return source.parent


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Write the repeat-matrix variants of a real data dictionary",
    )
    parser.add_argument("--source", required=True, help="Path to the source .xlsx")
    parser.add_argument("--out", default=str(DEFAULT_OUT), help="Where to write variants")
    parser.add_argument(
        "--csv-dir",
        default=None,
        help="Directory holding the dictionary's CSVs (default: beside the .xlsx)",
    )
    parser.add_argument(
        "--output-path",
        default=None,
        help="Where a build should put its zip (default: <out>/packages)",
    )
    parser.add_argument("--child", default=matrix_module.MATRIX_CHILD)
    parser.add_argument("--prefix", default=DEFAULT_PREFIX)
    parser.add_argument(
        "--clean",
        action="store_true",
        help="Delete the output directory first, so a removed cell leaves no stale file",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    source = Path(args.source).expanduser()
    out_dir = Path(args.out).expanduser()
    csv_dir = Path(args.csv_dir).expanduser() if args.csv_dir else _default_csv_dir(source)
    output_path = (
        Path(args.output_path).expanduser() if args.output_path else out_dir / "packages"
    )

    if args.clean and out_dir.exists():
        shutil.rmtree(out_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    try:
        variants = build(
            source,
            out_dir,
            csv_dir=csv_dir,
            output_path=output_path,
            child=args.child,
            prefix=args.prefix,
        )
    except VariantError as error:
        print(f"ERROR - {error}", file=sys.stderr)
        return 1

    print(f"Wrote {len(variants)} variants to {out_dir}")
    for variant in variants:
        print(f"  {variant.cell.key}  {variant.survey_id:20}  {variant.cell.describe()}")
    print(f"\nBuild one with:\n  python main.py --config {variants[0].config}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
